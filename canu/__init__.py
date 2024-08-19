import os, time, yaml, pickle, tempfile
from pathlib import Path
import streamlit as st
import streamlit_authenticator as stauth
import openai
import mysql.connector
from PIL import Image
import boto3
import xlrd
import openpyxl

class Container():
    def __init__(self, role, blocks, show_code_block=True, show_download_button=True):
        self.container = st.empty()
        self.role = role
        self.blocks = blocks
        self.show_code_block = show_code_block
        self.show_download_button = show_download_button
        self.code_interpreter_files = {}

    def _write_blocks(self):
        avatar = None
        if self.role == "assistant" and "assistant_avatar" in st.session_state:
            avatar = Image.open(st.session_state.assistant_avatar)
        with st.chat_message(self.role, avatar=avatar):
            for block in self.blocks:
                if block['type'] == 'text':
                    st.write(block['content'], unsafe_allow_html=True)
                elif block['type'] == 'code' and self.show_code_block:
                    st.code(block['content'])
                elif block['type'] == 'image':
                    st.image(block['content'])
            if self.code_interpreter_files and self.show_download_button:
                for filename, content in self.code_interpreter_files.items():
                    if filename.endswith('.csv'):
                        mime = "text/csv"
                    elif filename.endswith('.png'):
                        mime = "image/png"
                    else:
                        mime = "text/plain"
                    st.download_button(
                        label=f"{filename}",
                        data=content,
                        file_name=filename,
                        mime=mime,
                        key=f"download_button_{st.session_state.download_button_key}"
                    )
                    st.session_state.download_button_key += 1

    def get_content(self):
        content = []
        for block in self.blocks:
            if block['type'] == 'text':
                content.append({"type": "text", "text": block['content']})
        return content

    def write_blocks(self, stream=False):
        if stream:
            with self.container:
                self._write_blocks()
        else:
            self._write_blocks()

class EventHandler(openai.AssistantEventHandler):
    def __init__(self, container=None, show_quotation_marks=True, show_code_block=True, show_download_button=True):
        super().__init__()
        self.container = container
        self.redundant = container is not None
        self.show_quotation_marks = show_quotation_marks
        self.show_code_block = show_code_block
        self.show_download_button = show_download_button

    def on_text_delta(self, delta, snapshot):
        if self.container is None:
            self.container = Container("assistant", [], show_code_block=self.show_code_block, show_download_button=self.show_download_button)
        if not self.container.blocks or self.container.blocks[-1]['type'] != 'text':
            self.container.blocks.append({'type': 'text', 'content': ""})
        if delta.annotations is not None:
            for annotation in delta.annotations:
                if annotation.type == "file_citation":
                    file = st.session_state.client.files.retrieve(annotation.file_citation.file_id)
                    if self.show_quotation_marks:
                        delta.value = delta.value.replace(annotation.text, f"""<a href="#" title="{file.filename}">[❞]</a>""")
                    else:
                        delta.value = delta.value.replace(annotation.text, "")
                elif annotation.type == "file_path":
                    file = st.session_state.client.files.retrieve(annotation.file_path.file_id)
                    content = st.session_state.client.files.content(file.id)
                    filename = os.path.basename(file.filename)
                    self.container.code_interpreter_files[filename] = content.read()
        if delta.value is not None:
            self.container.blocks[-1]["content"] += delta.value
        self.container.write_blocks(stream=True)

    def on_image_file_done(self, image_file):
        if self.container is None:
            self.container = Container("assistant", [], show_code_block=self.show_code_block, show_download_button=self.show_download_button)
        if not self.container.blocks or self.container.blocks[-1]['type'] != 'image':
            self.container.blocks.append({'type': 'image', 'content': ""})
        image_data = st.session_state.client.files.content(image_file.file_id)
        image_data_bytes = image_data.read()
        self.container.blocks[-1]["content"] = image_data_bytes
        self.container.write_blocks(stream=True)

    def on_tool_call_delta(self, delta, snapshot):
        if delta.type == "function":
            pass
        elif delta.type == "code_interpreter":
            if self.container is None:
                self.container = Container("assistant", [], show_code_block=self.show_code_block, show_download_button=self.show_download_button)
            if delta.code_interpreter.input:
                if not self.container.blocks or self.container.blocks[-1]['type'] != 'code':
                    self.container.blocks.append({'type': 'code', 'content': ""})
                self.container.blocks[-1]["content"] += delta.code_interpreter.input
            self.container.write_blocks(stream=True)

    def submit_tool_outputs(self, tool_outputs, run_id):
        with st.session_state.client.beta.threads.runs.submit_tool_outputs_stream(
            thread_id=self.current_run.thread_id,
            run_id=self.current_run.id,
            tool_outputs=tool_outputs,
            event_handler=EventHandler(self.container),
        ) as stream:
            stream.until_done()

    def on_end(self):
        if self.container is not None and not self.redundant:
            st.session_state.containers.append(self.container)

    def on_event(self, event):
        if event.event == 'thread.run.requires_action':
            run_id = event.data.id
            self.handle_requires_action(event.data, run_id)

def authenticate():
    """
    Create an authenticator object based on the authentication method 
    specified in the config.yaml file. Currently, the supported methods are 
    'YAML' and 'MYSQL'.
    """
    def from_yaml():
        with open("./auth.yaml") as f:
            config = yaml.load(f, Loader=yaml.loader.SafeLoader)
        authenticator = stauth.Authenticate(
            config['credentials'],
            config['cookie']['cookie_name'],
            config['cookie']['cookie_key'],
            config['cookie']['cookie_expiry_days'],
        )
        return authenticator
    
    def from_mysql():
        try:
            connection = mysql.connector.connect(
                user=st.session_state.config['authentication']['user'],
                password=st.session_state.config['authentication']['password'],
                host=st.session_state.config['authentication']['host'],
                database=st.session_state.config['authentication']['database']
            )
            if connection.is_connected():
                cursor = connection.cursor(dictionary=True)
                cursor.execute(f"SELECT username, name, password FROM {st.session_state.config['authentication']['table']}")
                result = cursor.fetchall()
                credentials = {'usernames': {row['username']: {'name': row['name'], 'password': row['password']} for row in result}}
                cursor.close()
        except mysql.connector.Error as e:
            st.error(f"Error connecting to MySQL database: {e}")
        finally:
            if connection.is_connected():
                connection.close()
        authenticator = stauth.Authenticate(credentials, '', '', 0)
        return authenticator

    if "page" not in st.session_state:
        st.session_state.page = "login"

    if "authenticator" not in st.session_state:
        method = st.session_state.config['authentication']['method']
        if method == "YAML":
            authenticator = from_yaml()
        elif method == "MYSQL":
            authenticator = from_mysql()
        else:
            raise ValueError(f"Invalid authentication method: {method}")
        st.session_state.authenticator = authenticator

def add_message(role, content):
    create_message(role, content)
    st.session_state.containers.append(
        Container(role, [{'type': 'text', 'content': content}])
    )

def write_stream(event_handler=None, show_quotation_marks=True, show_code_block=True, show_download_button=True):
    if event_handler is None:
        event_handler = EventHandler(show_quotation_marks=show_quotation_marks, show_code_block=show_code_block, show_download_button=show_download_button)
    if not is_thread_locked():
        with st.session_state.client.beta.threads.runs.stream(
            thread_id=st.session_state.thread.id,
            assistant_id=st.session_state.assistant.id,
            event_handler=event_handler,
        ) as stream:
            stream.until_done()

def show_login_page():
    labels = {
        'Form name': {'English': 'Login', 'Korean': '로그인', 'Spanish': 'Inicio de sesión', 'Japanese': 'ログイン'},
        'Username': {'English': 'Username', 'Korean': '아이디', 'Spanish': 'Nombre de usuario', 'Japanese': 'ユーザー名'},
        'Password': {'English': 'Password', 'Korean': '비밀번호', 'Spanish': 'Contraseña', 'Japanese': 'パスワード'},
        'Login': {'English': 'Login', 'Korean': '로그인', 'Spanish': 'Iniciar sesión', 'Japanese': 'ログイン'},
        'Incorrect credential': {'English': 'The ID or password is incorrect.', 'Korean': '아이디 또는 비밀번호가 잘못되었습니다.', 'Spanish': 'El ID o la contraseña son incorrectos.', 'Japanese': 'IDまたはパスワードが間違っています.'},
    }
    mapping = {'한국어': 'Korean', 'English': 'English', 'Español': 'Spanish', '日本語': 'Japanese'}
    language = st.selectbox("NONE", ["한국어", "English", "Español", "日本語"], label_visibility="collapsed")
    language = mapping[language]
    st.session_state.language = language
    st.session_state.name, st.session_state.authentication_status, st.session_state.username = st.session_state.authenticator.login(location="main", fields={'Form name': labels['Form name'][st.session_state.language], 'Username': labels['Username'][st.session_state.language], 'Password': labels['Password'][st.session_state.language], 'Login': labels['Login'][st.session_state.language]})
    if st.session_state.authentication_status:
        st.session_state.page = "chatbot"
        st.rerun()
    elif st.session_state.authentication_status is False:
        st.error(labels['Incorrect credential'][st.session_state.language])
    elif st.session_state.authentication_status is None:
        pass

def show_profile_page():
    """
    Show the profile page.
    """
    def update_yaml():
        with open("./auth.yaml", 'w', encoding="utf-8-sig") as f:
            data = {
                'cookie': {
                    'cookie_name': st.session_state.authenticator.cookie_handler.cookie_name,
                    'cookie_key': st.session_state.authenticator.cookie_handler.cookie_key,
                    'cookie_expiry_days': st.session_state.authenticator.cookie_handler.cookie_expiry_days,
                },
                'credentials': st.session_state.authenticator.authentication_handler.credentials
            }
            
            yaml.dump(data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

    def update_mysql():
        credentials = st.session_state.authenticator.authentication_handler.credentials
        password = credentials['usernames'][st.session_state.username]['password']
        try:
            connection = mysql.connector.connect(
                user=st.session_state.config['authentication']['user'],
                password=st.session_state.config['authentication']['password'],
                host=st.session_state.config['authentication']['host'],
                database=st.session_state.config['authentication']['database']
            )
            if connection.is_connected():
                cursor = connection.cursor()
                cursor.execute(f"UPDATE {st.session_state.config['authentication']['table']} SET password = %s WHERE username = %s", (password, st.session_state.username))
                connection.commit()
                cursor.close()
        except mysql.connector.Error as e:
            st.error(f"Error connecting to MySQL database: {e}")
        finally:
            if connection.is_connected():
                connection.close()

    labels = {
        'Form name': {'English': 'Reset password', 'Korean': '비밀번호 변경', 'Spanish': 'Restablecer contraseña', 'Japanese': 'パスワードをリセットする'},
        'Current password': {'English': 'Current password', 'Korean': '현재 비밀번호', 'Spanish': 'Contraseña actual', 'Japanese': '現在のパスワード'},
        'New password': {'English': 'New password', 'Korean': '새로운 비밀번호', 'Spanish': 'Nueva contraseña', 'Japanese': '新しいパスワード'},
        'Repeat password': {'English': 'Repeat password', 'Korean': '새로운 비밀번호 확인', 'Spanish': 'Repetir contraseña', 'Japanese': 'パスワードを繰り返す'},
        'Reset': {'English': 'Reset', 'Korean': '변경', 'Spanish': 'Restablecer', 'Japanese': 'リセット'},
        'Go back': {'English': 'Go back', 'Korean': '돌아가기', 'Spanish': 'Regresar', 'Japanese': '戻る'},
        'Change success': {'English': 'Password has been successfully changed.', 'Korean': '비밀번호가 성공적으로 변경되었습니다.', 'Spanish': 'La contraseña se ha cambiado correctamente.', 'Japanese': 'パスワードが正常に変更されました.'},
    }
    if "file_uploader_key" in st.session_state:
        uploaded_files = get_uploaded_files()
    if st.sidebar.button(labels['Go back'][st.session_state.language]):
        st.session_state.page = "chatbot"
        st.rerun()
    if st.session_state.authenticator.reset_password(st.session_state.username, fields={'Form name': labels['Form name'][st.session_state.language], 'Current password': labels['Current password'][st.session_state.language], 'New password': labels['New password'][st.session_state.language], 'Repeat password': labels['Repeat password'][st.session_state.language], 'Reset': labels['Reset'][st.session_state.language]}):
        st.success(labels['Change success'][st.session_state.language])
        time.sleep(3)
        method = st.session_state.config['authentication']['method']
        if method == "YAML":
            update_yaml()
        elif method == "MYSQL":
            update_mysql()
        st.session_state.page = "chatbot"
        st.rerun()

def get_uploaded_files():
    uploaded_files = st.sidebar.file_uploader(
        "NONE",
        accept_multiple_files=True,
        label_visibility="hidden",
        key=st.session_state.file_uploader_key
    )
    return uploaded_files

def show_history_page():
    """
    Manage the conversation history based on the storage method specified in 
    the config.yaml file. Currently, the supported methods are 'LOCAL' and 
    'S3'.
    """
    def from_local():
        if not os.path.isdir("./users"):
            os.mkdir("./users")
        if not os.path.isdir(f"./users/{st.session_state.username}"):
            os.mkdir(f"./users/{st.session_state.username}")
        st.header(labels['Current conversation'][st.session_state.language])
        with st.form(labels['Save conversation'][st.session_state.language], clear_on_submit=True):
            file_name = st.text_input(labels['Conversation name'][st.session_state.language])
            submitted = st.form_submit_button(labels['Save'][st.session_state.language])
            if submitted:
                data = []
                for container in st.session_state.containers:
                    data.append([container.role, container.blocks])
                with open(f"./users/{st.session_state.username}/{file_name}.pkl", 'wb') as f:
                    pickle.dump(data, f)
        st.header(labels['Past conversations'][st.session_state.language])
        files = os.listdir(f"./users/{st.session_state.username}")
        files = [x for x in files if x.endswith('.pkl')]
        options = [x.replace('.pkl', '') for x in files]
        option = st.selectbox(labels['Select conversation'][st.session_state.language], options)
        col1, col2 = st.columns((1, 6))
        with col1:
            if option is not None and st.button(labels['Load'][st.session_state.language]):
                delete_messages()
                st.session_state.containers = []
                with open(f"./users/{st.session_state.username}/{option}.pkl", 'rb') as f:
                    data = pickle.load(f)
                    for x in data:
                        role = x[0]
                        blocks = x[1]
                        container = Container(role, blocks)
                        st.session_state.containers.append(container)
                        create_message(role, container.get_content())
                st.session_state.page = "chatbot"
                st.rerun()
        with col2:
            if option is not None and st.button(labels['Delete'][st.session_state.language]):
                os.remove(f"./users/{st.session_state.username}/{option}.pkl")
                st.rerun()

    def from_s3():
        bucket = st.session_state.config['history']['bucket']
        s3 = boto3.client(
            's3',
            aws_access_key_id=st.session_state.config['history']['aws_access_key_id'],
            aws_secret_access_key=st.session_state.config['history']['aws_secret_access_key']
        )
        s3_folder = f"{st.session_state.config['history']['users_dir']}/{st.session_state.username}"
        st.header(labels['Current conversation'][st.session_state.language])
        with st.form(labels['Save conversation'][st.session_state.language], clear_on_submit=True):
            file_name = st.text_input(labels['Conversation name'][st.session_state.language])
            submitted = st.form_submit_button(labels['Save'][st.session_state.language])
            if submitted:
                data = []
                for container in st.session_state.containers:
                    data.append([container.role, container.blocks])
                with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                    temp_file_name = temp_file.name
                    with open(temp_file_name, 'wb') as f:
                        pickle.dump(data, f)
                    s3.upload_file(temp_file_name, bucket, f"{s3_folder}/{file_name}.pkl")
                    os.remove(temp_file_name)

        st.header(labels['Past conversations'][st.session_state.language])
        file_list = []
        folder_exists = s3.list_objects(Bucket=bucket, Prefix=f"{s3_folder}/")
        if 'Contents' in folder_exists:
            for obj in folder_exists['Contents']:
                file_list.append(os.path.basename(obj['Key']))
        files = [x for x in file_list if x.endswith('.pkl')]
        options = [x.replace('.pkl', '') for x in files]
        option = st.selectbox(labels['Select conversation'][st.session_state.language], options)
        col1, col2 = st.columns((1, 6))
        with col1:
            if option is not None and st.button(labels['Load'][st.session_state.language]):
                delete_messages()
                st.session_state.containers = []
                with tempfile.NamedTemporaryFile(delete=True) as temp_file:
                    temp_file_name = temp_file.name
                    s3.download_file(bucket, f"{s3_folder}/{option}.pkl", temp_file_name)
                    with open(temp_file_name, 'rb') as f:
                        data = pickle.load(f)
                        for x in data:
                            role = x[0]
                            blocks = x[1]
                            container = Container(role, blocks)
                            st.session_state.containers.append(container)
                            create_message(role, container.get_content())
                st.session_state.page = "chatbot"
                st.rerun()
        with col2:
            if option is not None and st.button(labels['Delete'][st.session_state.language]):
                s3.delete_object(Bucket=bucket, Key=f"{s3_folder}/{option}.pkl")
                st.rerun()

    labels = {
        'Go back': {'English': 'Go back', 'Korean': '돌아가기', 'Spanish': 'Regresar', 'Japanese': '戻る'},
        'Current conversation': {'English': 'Current conversation', 'Korean': '현재 대화', 'Spanish': 'Conversación actual', 'Japanese': '現在の会話'},
        'Save conversation': {'English': 'Save conversation', 'Korean': '대화 저장', 'Spanish': 'Guardar conversación', 'Japanese': '会話を保存する'},
        'Conversation name': {'English': 'Enter a name for the conversation to save.', 'Korean': '저장할 대화 이름을 입력하세요.', 'Spanish': 'Ingrese un nombre para la conversación a guardar.', 'Japanese': '保存する会話の名前を入力してください。'},
        'Save': {'English': 'Save', 'Korean': '저장', 'Spanish': 'Guardar', 'Japanese': '保存'},
        'Past conversations': {'English': 'Past conversations', 'Korean': '과거 대화', 'Spanish': 'Conversaciones pasadas', 'Japanese': '過去の会話'},
        'Select conversation': {'English': 'Select a conversation.', 'Korean': '대화를 선택해주세요.', 'Spanish': 'Seleccione una conversación.', 'Japanese': '会話を選択してください。'},
        'Load': {'English': 'Load', 'Korean': '불러오기', 'Spanish': 'Cargar', 'Japanese': 'ロード'},
        'Delete': {'English': 'Delete', 'Korean': '삭제하기', 'Spanish': 'Eliminar', 'Japanese': '削除'},
    }
    if "file_uploader_key" in st.session_state:
        uploaded_files = get_uploaded_files()
    if st.sidebar.button(labels['Go back'][st.session_state.language]):
        st.session_state.page = "chatbot"
        st.rerun()
    if st.session_state.config['history']['method'] == "LOCAL":
        from_local()
    elif st.session_state.config['history']['method'] == "S3":
        from_s3()
    else:
        raise ValueError(f"Invalid history storage method: {st.session_state.config['history']['method']}")

def handle_files():
    labels = {
        'Unsupported file type': {'English': 'Unsupported file type', 'Korean': '지원하지 않는 파일 형식', 'Spanish': 'Tipo de archivo no compatible', 'Japanese': 'サポートされていないファイル形式'},
        'Upload file': {'English': 'File Upload', 'Korean': '파일 업로드', 'Spanish': 'Subir archivo', 'Japanese': 'ファイルアップロード'},
        'Delete file': {'English': 'Delete file', 'Korean': '파일 삭제', 'Spanish': 'Eliminar archivo', 'Japanese': 'ファイル削除'},
    }
    supported_files = {
        "file_search": ['.c', '.cs', '.cpp', '.doc', '.docx', '.html', '.java', '.json', '.md', '.pdf', '.php', '.pptx', '.py', '.rb', '.texv', '.txt', '.css', '.js', '.sh', '.ts'],
        "code_interpreter": ['.c', '.cs', '.cpp', '.doc', '.docx', '.html', '.java', '.json', '.md', '.pdf', '.php', '.pptx', '.py', '.rb', '.tex', '.txt', '.css', '.js', '.sh', '.ts', '.csv', '.jpeg', '.jpg', '.gif', '.png', '.tar', '.xlsx', '.xml', '.zip']
    }

    uploaded_files = get_uploaded_files()

    for uploaded_file in uploaded_files:
        upload_id = uploaded_file.file_id
        file_name = uploaded_file.name
        if upload_id in st.session_state.upload_ids:
            continue
        _, file_extension = os.path.splitext(file_name)
        if file_extension not in supported_files["file_search"] and file_extension not in supported_files["code_interpreter"] and file_extension not in [".hwp", ".xls"]: 
            add_message("user", f"{labels['Unsupported file type'][st.session_state.language]}: `{file_name}`")
            print(f"Unsupported file type: {file_name}")
            st.session_state.file_uploader_key += 1
            st.rerun()
        with tempfile.TemporaryDirectory() as t:
            file_path = os.path.join(t, file_name)

            with open(file_path, "wb") as f:
                f.write(uploaded_file.getvalue())
            add_message("user", f"{labels['Upload file'][st.session_state.language]}: `{file_name}`")

            # Convert .xls to .xlsx.
            if file_name.endswith(".xls"):
                wb1 = xlrd.open_workbook(file_path)
                wb2 = openpyxl.Workbook()
                for sheet_name in wb1.sheet_names():
                    sheet1 = wb1.sheet_by_name(sheet_name)
                    sheet2 = wb2.create_sheet(title=sheet_name)
                    for row in range(sheet1.nrows):
                        for col in range(sheet1.ncols):
                            sheet2.cell(row=row+1, column=col+1, value=sheet1.cell_value(row, col))
                wb2.remove(wb2['Sheet'])
                file_name = file_name.replace(".xls", ".xlsx")
                file_path = file_path.replace(".xls", ".xlsx")
                wb2.save(file_path)
            # Convert .hwp to .html.
            elif file_name.endswith(".hwp"):
                with open(file_path, 'rb') as f1:
                    file_name = file_name.replace(".hwp", ".html")
                    file_path = file_path.replace(".hwp", ".html")
                    os.system(f"""hwp5html "{file_path.replace('.html', '.hwp')}" --output "{file_path}" --html""")

            if file_name.endswith(".jpg") or file_name.endswith(".png") or file_name.endswith(".jpeg"):
                file = st.session_state.client.files.create(file=Path(file_path), purpose="vision")
                content=[
                    {"type": "text", "text": f"{labels['Upload file'][st.session_state.language]}: `{file_name}`"},
                    {"type": "image_file", "image_file": {"file_id": file.id}}
                ]
                create_message("user", content)
            else:
                file = st.session_state.client.files.create(file=Path(file_path), purpose="assistants")
                tools = []
                if file_name.endswith(tuple(supported_files["file_search"])):
                    tools.append({"type": "file_search"})
                if file_name.endswith(tuple(supported_files["code_interpreter"])):
                    tools.append({"type": "code_interpreter"})
                if not tools:
                    tools.append({"type": "code_interpreter"})
                attachments = [{"file_id": file.id, "tools": tools}]
                content=[{"type": "text", "text": f"{labels['Upload file'][st.session_state.language]}: `{file_name}`"}]
                create_message("user", content, attachments)
            st.session_state.upload_ids[upload_id] = {'file_id': file.id, 'file_name': file_name}

    for upload_id, upload_data in list(st.session_state.upload_ids.items()):
        file_name = upload_data["file_name"]
        file_id = upload_data["file_id"]
        if upload_id not in [x.file_id for x in uploaded_files]:
            st.session_state.client.files.delete(file_id)
            add_message("user", f"{labels['Delete file'][st.session_state.language]}: `{file_name}`")
            del st.session_state.upload_ids[upload_id]

def list_messages():
    messages = st.session_state.client.beta.threads.messages.list(
        thread_id=st.session_state.thread.id
    )
    return messages

def delete_messages():
    messages = list_messages()
    for message in messages.data:
        deleted_message = st.session_state.client.beta.threads.messages.delete(
            thread_id=st.session_state.thread.id,
            message_id=message.id
        )

def create_message(role, content, attachments=None):
    """
    Create a message and add it to the thread.
    """
    if not is_thread_locked():
        st.session_state.client.beta.threads.messages.create(
            thread_id=st.session_state.thread.id,
            role=role,
            content=content,
            attachments=attachments
        )

def delete_files():
    """
    Delete all files uploaded to OpenAI.
    """
    for upload_id, upload_data in st.session_state.upload_ids.items():
        st.session_state.client.files.delete(upload_data["file_id"])

def list_runs(limit=100):
    """
    Returns a list of runs belonging to the thread.
    """
    runs = st.session_state.client.beta.threads.runs.list(
        thread_id=st.session_state.thread.id,
        limit=limit
    )
    return runs

def is_thread_locked():
    """
    Returns whether the thread is locked.
    """
    return len([x for x in list_runs().data if x.status in ["queued", "in_progress"]]) > 0

def get_config():
    """
    Get the configuration from the config.yaml file.
    """
    if "config" not in st.session_state:
        with open("./config.yaml") as f:
            st.session_state.config = yaml.load(f, Loader=yaml.loader.SafeLoader)
